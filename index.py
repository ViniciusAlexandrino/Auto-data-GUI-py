from tkinter import *
from tkinter.ttk import Combobox
import tkinter as tk
from tkinter import messagebox
import openpyxl
from openpyxl import Workbook
import pathlib

# Inicialização da janela principal do Tkinter
root = Tk()
root.title("Auto dados")
root.geometry('700x400+300+200')
root.resizable(False, False)
root.configure(bg="#326273")

# Verificação se o arquivo Excel já existe
file_path = pathlib.Path('Backend_data.xlsx')
if file_path.exists():
    workbook = openpyxl.load_workbook(file_path)  # Carrega o workbook existente
else:
    workbook = Workbook()  # Cria um novo workbook
    sheet = workbook.active
    # Cria os cabeçalhos na primeira linha
    sheet['A1'] = "Nome Completo"
    sheet['B1'] = "Telefone"
    sheet['C1'] = "Idade"
    sheet['D1'] = "Endereço"
    workbook.save('Backend_data.xlsx')  # Salva o novo workbook

# Função para submeter os dados
def submit():
    # Obtém os valores dos campos de entrada
    nome = nameValue.get()
    telefone = contactValue.get()
    idade = ageValue.get()
    endereço = adressEntry.get(1.0, END).strip()

    # Carrega o workbook e obtém a planilha ativa
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    # Encontra a próxima linha disponível para adicionar os dados
    next_row = sheet.max_row + 1
    sheet.cell(column=1, row=next_row, value=nome)
    sheet.cell(column=2, row=next_row, value=telefone)
    sheet.cell(column=3, row=next_row, value=idade)
    sheet.cell(column=4, row=next_row, value=endereço)

    # Salva o workbook com os novos dados
    workbook.save(file_path)

    # Exibe uma mensagem de sucesso
    messagebox.showinfo('info', 'Dados adicionados')

    # Limpa os campos de entrada
    nameValue.set('')
    contactValue.set('')
    ageValue.set('')
    adressEntry.delete(1.0, END)

# Função para limpar os campos de entrada
def clear():
    nameValue.set('')
    contactValue.set('')
    ageValue.set('')
    adressEntry.delete(1.0, END)

# Cabeçalho
Label(root, text="Insira seus dados:", font="arial 13", bg="#326273", fg="#fff").place(x=20, y=20)

# Labels
Label(root, text='Nome', font=23, bg="#326273", fg="#fff").place(x=50, y=100)
Label(root, text='Telefone', font=23, bg="#326273", fg="#fff").place(x=50, y=150)
Label(root, text='Idade', font=23, bg="#326273", fg="#fff").place(x=50, y=200)
Label(root, text='Endereço', font=23, bg="#326273", fg="#fff").place(x=50, y=250)

# Campos de entrada
nameValue = StringVar()
contactValue = StringVar()
ageValue = StringVar()

nameEntry = Entry(root, textvariable=nameValue, width=45, bd=2, font=20)
contactEntry = Entry(root, textvariable=contactValue, width=45, bd=2, font=20)
ageEntry = Entry(root, textvariable=ageValue, width=15, bd=2, font=20)

# Campo de entrada para o endereço
adressEntry = Text(root, width=50, height=4, bd=2)

# Posicionamento dos campos de entrada
nameEntry.place(x=200, y=100)
contactEntry.place(x=200, y=150)
ageEntry.place(x=200, y=200)
adressEntry.place(x=200, y=250)

# Botões
Button(root, text="Enviar", bg="#326273", fg="white", width=15, height=2, command=submit).place(x=200, y=350)
Button(root, text="Limpar", bg="#326273", fg="white", width=15, height=2, command=clear).place(x=340, y=350)
Button(root, text="Sair", bg="#326273", fg="white", width=15, height=2, command=lambda: root.destroy()).place(x=480, y=350)

# Inicia o loop principal da interface gráfica
root.mainloop()
